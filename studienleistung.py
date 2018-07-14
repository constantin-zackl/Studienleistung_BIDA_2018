#!/usr/bin/python3

# Author: Constantin Zackl, Student @ TH-Bingen, 2018
# Licence: under construction

import sys
import re
from openpyxl import Workbook
from openpyxl import worksheet


# Erklaerung zu den Variablen
# proteomdat: Dateiname der Datei die das Proteom enthält (Fasta Datei)
# ausexl: Name der Excel Ausgabedatei
# proaccnr: Liste der Prosite Accessionnummern die der Nutzer eingegeben hat
# regexlist: Dictionary in dem die gefundenen Python Pattern mit den Accessionnummern verknüpft sind
# prositedatname: Name der Prosite Datenbank, voreingestellt
# found_domaene: Anzahl der gefundenen Domaenen, wird genutzt um bei keinem Fund keine Datei zu erstellen
# wb: Enthaelt Excel Tabelle
# zeile: speichert zeile beim datei einlesen
# z: zwischenspeicher für pattern bei der verarbeitung
# p(x): einzelne Pattern für das re modul
# x, y, s: zwischenspeicher


# Klassen
# ----------------------------------------------------------------------------------------

class Sequence:

    def __init__(self, accnr):
        self.accnr = accnr
        self.sequence = ""
        self.len = 0


class expression:

    def __init__(self):
        self.regexp = ""
        self.pattern = ""

    def assignpattern(self, pattern):
        self.pattern = str(pattern)

    def translatetoregexp(self):
        if (self.pattern == ""):
            return (-1)
        else:
            self.regexpr = self.pattern

            return (self.regexpr)


# ----------------------------------------------------------------------------------------

# Initialisierung
proteomdat = ""
ausexcl = ""
proaccnr = []
prositedatname = "prosite.dat"
regexlist = {}
found_domaene = 0

wb = Workbook()  # Erzeugt neue Tabelle

# -----------------------------------------------------------------------------
# Eingabe - Übergabe der Parameter im Terminal
# -----------------------------------------------------------------------------
# Einlesen und überprüfen der Proteom Datei im Fasta Format
try:
    proteomdat = sys.argv[1]
    if (not (re.search(".fasta", proteomdat))):
        print("Fehler, ungültiger Dateiname der Proteom Datei")
        sys.exit()
    datprot = open(proteomdat, "r")

except:
    print("Fehler beim öffnen de Proteom Datei. Bitte gültigen Dateinamen übergeben.")
    sys.exit()

# Einlesen und Überprüfen des gewünschten Dateinamen für die Ausgabe Excel Tabelle

try:
    ausexcl = str(sys.argv[2])
    ausexcl += ".xlsx"
except:
    print("Ungültiger Dateiname der Ausgabe Excel Datei")
    sys.exit()

# Einlesen der Accession Numbers wobei beliebig viele eingegeben werden können
# Alle Übergebenen Acession Nummern werden einer Liste angefügt

i = 3
while (True):
    try:
        proaccnr.append(str(sys.argv[i]))
        i += 1
    except:
        break

# Wenn keine Accession Nummern übergeben wurde, abbrechen
if (len(proaccnr) == 0):
    print("Bitte Prosite Accession Number übergeben")
    sys.exit()

# Prosite Datenbank öffnen
try:
    prositedat = open(prositedatname, "r")
except:
    print("Keine Prosite Datenbank vorhanden oder Dateinamen ungültig")
    sys.exit()

accnrlen = len(proaccnr)

# -----------------------------------------------------------------------------
# Verarbeitung
# -----------------------------------------------------------------------------

# Auslesen der Prosite Pattern aus der Prosite Datenbank mithilfe der Accnr

zeile = prositedat.readline()
x = re.compile("AC\s+(PS[0-9]{5})")
v = re.compile("PA\s+(.+)")
c = -1
allfound = False
while (zeile and not allfound):
    # Suchen der Prosite Patterns

    # Suche nach der Passenden Accessionnumber
    y = re.search(x, zeile)
    if (y):
        for m in proaccnr:
            n = y.group(1)
            if (y.group(1) == m):
                c = 0

    # Die Folgezeilen lesen, um das Pattern zu finden
    w = re.match(v, zeile)

    # Wenn die richtige Zeile gefunden ist und diese zu der Accessionnumber
    # passt, dann wird das Pattern in einen regulären Ausdruck uebersetzt
    # und in ein Dictionary gespeichert in dem die Accessionnummern mit
    # den Python Pattern verknüpft sind
    if (w and c == 0):
        z = w.group(1)

        p1 = re.compile("-")
        p2 = re.compile("{")
        p3 = re.compile("}")
        p4 = re.compile("\(")
        p5 = re.compile("\)")
        p6 = re.compile("x")
        p7 = re.compile("\.")

        z = p1.sub("", z)
        z = p2.sub("[^", z)
        z = p3.sub("]", z)
        z = p4.sub("{", z)
        z = p5.sub("}", z)
        z = p6.sub("[A-Z]", z)
        z = p7.sub("", z)
        regexlist[n] = z
        c = -1
    # Wenn alle Pattern gefunden wurden, dann abbrechen
    if (len(regexlist) == accnrlen):
        allfound = True
    zeile = prositedat.readline()

# die Liste der gültigen Accession Nummern Updaten
for m in proaccnr:
    if m not in regexlist:
        print("Die Accessionnummer", m, "existiert nicht in der Prosite Datenbank")
        proaccnr.remove(m)

accnrlen = len(proaccnr)
print(regexlist)
print(proaccnr)
# Verarbeitung inklusive Ausgabe = schreiben in Tabelle

for i in range(accnrlen):  # Anzahl Accnr = Anzahl der Tabellenblättter
    # Am Anfang jedes durchlaufs erstmal Tabellenblatt erzeugen
    if (i == 0):
        ws = wb.active
        ws.title = proaccnr[i]
    else:
        ws = wb.create_sheet(title=proaccnr[i])
    # Header jedes Tabellenblatts schreiben
    ws["A1"] = "Accessionnumber"
    ws["B1"] = "Anfangsposition Domäne"
    ws["C1"] = "Sequenz Domäne"

# Jetzt beginnt die eigentliche Verarbeitung - Einfügen der gefundenen werte
# mit übersetztem Pattern in Fasta Datei suchen und in Tabelle schreiben
# Jeder schleifendurchlauf entspricht dem durchsuchen der Fasta Proteom Datei mit einer
# prosite Accnr. Dh. durchsuchen der FASTA datei mit der RE

zeile = datprot.readline()
c = -1
sequence = ""
while (zeile):
    if (zeile[0] == ">"):
        p = re.compile(">.{3}(.{6})")
        m = re.search(p, zeile)
        accnr = m.group(1)
        c += 1
        if (sequence != ""):
            sequence = re.sub("\n", "", sequence)
            # Ausgabe in Excel Tabelle#
            for i in range(accnrlen):

                data = [accnr, "", ""]
                p = re.compile(regexlist[proaccnr[i]])
                s = p.finditer(sequence)
                for m in (s):
                    data[2] = m.string[m.start():m.end()]
                    data[1] = m.start()
                    wb[proaccnr[i]].append(data)
                    found_domaene += 1

            c = -1
        sequence = ""
    elif (c == 0):
        sequence += zeile

    zeile = datprot.readline()
# Ausgabe in Excel Tabelle#
for i in range(accnrlen):

    data = [accnr, "", ""]
    p = re.compile(regexlist[proaccnr[i]])
    s = p.finditer(sequence)
    for m in (s):
        data[2] = m.string[m.start():m.end()]
        data[1] = m.start()
        wb[proaccnr[i]].append(data)
        found_domaene += 1

# Dateien schließen und Excel Tabelle Speichern, wenn etwas gefunden wurde
if (found_domaene > 0):
    wb.save(filename=ausexcl)  # bedingt, nur wenn etwas gefunden wurde
prositedat.close()
datprot.close()
